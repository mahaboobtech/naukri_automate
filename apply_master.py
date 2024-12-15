import openpyxl
import os
import logging
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import sec
from geminipost import interact_with_gemini


# Define logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filename="application_log.log",  # Log file name
    filemode="a"  # Append mode
)

# Define the file to store last applied job index
last_applied_file = "last_applied_job.txt"


def log_status_and_index(index, status, title):
    """Log the current application status and index."""
    logging.info(f"Index: {index}, Job Title: '{title}', Status: {status}")


# Function to read the last applied job index
def get_last_applied_index():
    if os.path.exists(last_applied_file):
        with open(last_applied_file, "r") as file:
            return int(file.read().strip())
    return 0


# Function to save the last applied job index
def save_last_applied_index(index):
    with open(last_applied_file, "w") as file:
        file.write(str(index))


# Load the prepared job links from the Excel file
input_wb = openpyxl.load_workbook("Naukri_Job_Listings.xlsx")
input_ws = input_wb.active

# Initialize WebDriver
driver = webdriver.Chrome()

# Login credentials
email = sec.email
password = sec.passw

count = 0
# File paths
final_file = "final_output.xlsx"
applied_file = "Successfully_Applied_Jobs.xlsx"
failed_file = "Failed_Jobs.xlsx"

# Load or create final_output.xlsx
if os.path.exists(final_file):
    final_wb = openpyxl.load_workbook(final_file)
    final_ws = final_wb.active
else:
    final_wb = openpyxl.Workbook()
    final_ws = final_wb.active
    final_ws.append(["Job Title", "Link", "Experience", "Status", "External Link or Notice"])

# Login to Naukri
try:
    driver.get("https://www.naukri.com/nlogin/login")
    time.sleep(5)
    email_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "usernameField"))
    )
    password_input = driver.find_element(By.ID, "passwordField")
    email_input.send_keys(email)
    password_input.send_keys(password)
    password_input.send_keys(Keys.RETURN)
    logging.info("Logged in successfully.")
    time.sleep(5)
except Exception as e:
    logging.error(f"Login failed: {e}")
    driver.quit()

# Get the last applied job index
last_applied_index = get_last_applied_index()

# Define the functions for interacting with the chatbot
def gemini_input():
    try:
        # Locate all elements with class name "botMsg"
        questions = driver.find_elements(By.CLASS_NAME, "botMsg")
        if questions:
            # Extract the last question's text
            last_question = questions[-1].text
            logging.info(f"Extracted last question: {last_question}")

            # Interact with Gemini to generate a response
            answer = interact_with_gemini(f"{last_question}, provide an appropriate response.")

            # Find the input field and enter the answer
            input_field = driver.find_element(By.CLASS_NAME, "textArea")
            input_field.send_keys(answer)
            logging.info(f"Entered answer: {answer}")

            # Find and click the save/submit button
            save_button = driver.find_element(By.CLASS_NAME, "sendMsg")
            save_button.click()
            logging.info("Answer submitted.")
        else:
            logging.warning("No questions found with class 'botMsg'.")
    except Exception as e:
        logging.error(f"Error in gemini_input(): {e}")


def gemini_single():
    try:
        question = driver.find_element(By.CLASS_NAME, "botMsg").text
        options = driver.find_elements(By.XPATH, "//div[@class='ssrc__radio-btn-container']/input[@type='radio']")
        option_texts = [option.get_attribute("value") for option in options]
        logging.info(f"Extracted question: {question}")
        logging.info(f"Extracted options: {option_texts}")

        selected_option = interact_with_gemini(f"{question}, Options: {option_texts}, select the best option.")
        for option in options:
            if option.get_attribute("value") == selected_option:
                # Find the label associated with the radio button using the 'for' attribute
                label = driver.find_element(By.XPATH, f"//label[@for='{option.get_attribute('id')}']")
                # Click the label, which will select the radio button
                label.click()
                logging.info(f"Selected option: {selected_option}")
                save_button = driver.find_element(By.CLASS_NAME, "sendMsg")
                save_button.click()
                logging.info("Clicked the 'Save' button.")
                break

        save_button = driver.find_element(By.CLASS_NAME, "sendMsg")
        save_button.click()
        logging.info("Option submitted.")
    except Exception as e:
        logging.error(f"Error in gemini_single(): {e}")


# def gemini_agree():
#     try:
#         radio_button = driver.find_element(By.XPATH, "//div[@class='ssrc__radio-btn-container']/input[@type='radio']")
#         radio_button.click()
#         logging.info("Single radio button clicked.")

#         save_button = driver.find_element(By.CLASS_NAME, "sendMsg")
#         save_button.click()
#         logging.info("Agreement submitted.")
#     except Exception as e:
#         logging.error(f"Error in gemini_agree(): {e}")


try:
    # Loop through each job link starting from the last applied job
    for index, row in enumerate(input_ws.iter_rows(min_row=2, values_only=True), start=1):
        if index <= last_applied_index:
            continue

        title, link, experience = row

        # Open the job link
        try:
            driver.get(link)
            time.sleep(3)
        except Exception as e:
            logging.error(f"Failed to open job link '{title}': {e}")
            final_ws.append([title, link, experience, "Failed", "Link not accessible"])
            continue

        try:
            # Check for already applied status
            already_applied = driver.find_elements(By.ID, "already-applied")
            if already_applied:
                status = "Already Applied"
                logging.info(f"Job '{title}' is already applied.")
                final_ws.append([title, link, experience, status, ""])
                save_last_applied_index(index)
                log_status_and_index(index, status, title)
                continue

            # Apply for the job
            apply_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "apply-button"))
            )
            apply_button.click()
            time.sleep(5)  # Delay after pressing the "Apply" button to wait for any banners or popups

            # Check for error banner
            try:
                error_banner = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "styles_user-msg__YLRsE"))
                )
                if error_banner:
                    status = "Server Error"
                    logging.info(f"Server error detected for job '{title}'.")
                    final_ws.append([title, link, experience, status, ""])
                    save_last_applied_index(index)
                    log_status_and_index(index, status, title)
                    continue
            except Exception:
                pass

            while driver.find_elements(By.CLASS_NAME, "chatbot_Drawer"):  # Loop only while the chatbot drawer is present
                try:
                    count += 1
                    chatbot_area = driver.find_elements(By.CLASS_NAME, "chatbot_Drawer")
                    if chatbot_area:
                        logging.info("Clicking on the chatbot drawer to divert focus.")
                        chatbot_area[0].click()  
                    time.sleep(1)
                    logging.info("Chatbot detected, interacting with it. count : ",count)

                    # Check for single radio button selection
                    single_radio_selection = driver.find_elements(By.CLASS_NAME, "singleselect-radiobutton")
                    if single_radio_selection:
                        logging.info("Single radio button selection detected, executing gemini_single().")
                        gemini_single()
                    
                    # Check for input container
                    input_container = driver.find_elements(By.XPATH, "//*[contains(@class, 'chatbot_InputContainer') and not(contains(@class, 'inputContainer-focus'))]")
                    if input_container:
                        logging.info("Input container detected, executing gemini_input().")
                        gemini_input()
                    if count >= 20:
                        count = 0
                        break


                    # Check for agreement logic
                    # radio_buttons = driver.find_elements(By.CLASS_NAME, "singleselect-radiobutton")
                    # if len(radio_buttons) == 1:
                    #     logging.info("Single radio button detected, executing gemini_agree().")
                    #     gemini_agree()

                except Exception as e:
                    logging.error(f"An error occurred during chatbot interaction: {e}")

                    # Attempt to close chatbot if cross icon is present during exception
                    try:
                        cross_icon = driver.find_elements(By.CLASS_NAME, "chatBot-ic-cross")
                        if cross_icon:
                            cross_icon[0].click()
                            logging.info("Chatbot closed via cross icon during exception.")
                            break
                    except Exception as sub_e:
                        logging.error(f"Error closing chatbot during exception: {sub_e}")

                    final_ws.append([title, link, experience, status, ""])
                    save_last_applied_index(index)
                    log_status_and_index(index, status, title)
                    continue

            count = 0
            status_banner_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CLASS_NAME, "apply-message"))
            )
            
            # Check if the element contains the expected text
            if status_banner_element and "You have successfully applied to" in status_banner_element.text:
                stauts_banner = "Successfully Applied"
            else:
                stauts_banner = "Not Applied need to fill"
            logging.info(f"{stauts_banner} for job: {title}")
            final_ws.append([title, link, experience, status, ""])
            save_last_applied_index(index)
            log_status_and_index(index, status, title)      
                    

        except Exception as e:
            status = "Failed"
            logging.warning(f"Failed to apply for job '{title}': {e}")
            log_status_and_index(index, status, title)
            try:
                company_site_button = driver.find_element(By.ID, "company-site-button")
                company_site_button.click()
                time.sleep(2)

                driver.switch_to.window(driver.window_handles[1])
                external_url = driver.current_url
                status = "Redirected to Company Site"
                logging.info(f"Redirected to company site for job '{title}': {external_url}")
                final_ws.append([title, link, experience, status, external_url])
                save_last_applied_index(index)
                log_status_and_index(index, status, title)

                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            except Exception as sub_e:
                status = "Failed"
                logging.error(f"Failed to retrieve company site link for '{title}': {sub_e}")
                final_ws.append([title, link, experience, status, "No link found"])
                save_last_applied_index(index)
                log_status_and_index(index, status, title)

finally:
    final_wb.save(final_file)
    logging.info(f"Saved data to {final_file}")
    driver.quit()
